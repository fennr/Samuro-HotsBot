""""
Samuro Bot

Автор: *fennr*
github: https://github.com/fennr/Samuro-HotsBot

Бот для сообществ по игре Heroes of the Storm

"""

import os
from discord import Embed, Member, File
from discord.ext import commands
from utils import check
from enum import Enum
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from utils.classes import Const
from utils import exceptions, library
from utils.classes.Const import config


class League(Enum):
    Bronze = "Бронза"
    Silver = "Серебро"
    Gold = "Золото"
    Platinum = "Платина"
    Diamond = "Алмаз"
    Master = "Мастер"
    Grandmaster = "Грандмастер"

class Stats(commands.Cog, name="Stats"):
    """
    — Просмотр таблиц лидеров
    """
    def __init__(self, bot):
        self.bot = bot

    @commands.group(name="top")
    async def top(self, ctx):
        if ctx.invoked_subcommand is None:
            pass

    @top.command(name="excel")
    @check.is_samuro_dev()
    async def top_excel(self, ctx):
        headings = ['id', 'guild_id', 'Победы', 'Поражения', 'Очки', 'Батлтег']
        filepath = 'UserStats.xlsx'
        con, cur = library.get.con_cur()
        guild_id = library.get.guild_id(ctx)
        select = Const.selects.US
        cur.execute(select, (guild_id, ))
        data = cur.fetchall()
        cur.close()
        # Создание документа, страницы
        wb = openpyxl.Workbook()
        ws = wb.active
        # Spreadsheet row and column indexes start at 1
        # so we use "start = 1" in enumerate so
        # we don't need to add 1 to the indexes.
        for colno, heading in enumerate(headings, start=1):
            c = ws.cell(row=1, column=colno)
            c.font = Font(bold=True)
            c.value = heading
        # This time we use "start = 2" to skip the heading row.
        for rowno, row in enumerate(data, start=2):
            for colno, cell_value in enumerate(row, start=1):
                ws.cell(row=rowno, column=colno).value = cell_value
        # Выравнивание длины колонок
        dim_holder = DimensionHolder(worksheet=ws)
        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
        ws.column_dimensions = dim_holder
        # сохранение, вывод, удаление файла
        wb.save(filepath)
        await ctx.author.send(file=File(filepath))
        os.remove(filepath)
        await ctx.send("Файл отправлен личным сообщением")

    @top.command(name="mmr")
    async def top_mmr(self, ctx, league_type="Мастер", count=10):
        """
        - Лидеры по ммр
        """
        try:
            league = League(league_type)
        except Exception:
            raise exceptions.WrongLeague
        con, cur = library.get.con_cur()
        guild_id = library.get.guild_id(ctx)
        select = Const.selects.PlayersLeague
        cur.execute(select, (league.name, count))
        records = cur.fetchall()
        embed = Embed(
            title=f"Таблица лиги {league.value}",
            color=config.info
        )
        value = ""
        for i, record in enumerate(records):
            value += f"{i + 1}. {library.get.mention(record.id)} (mmr: {record.mmr})\n"
        embed.add_field(
            name=f"Топ {count} игроков",
            value=value
        )
        await ctx.send(embed=embed)

    @top.command(name="wins")
    async def top_wins(self, ctx, count=10):
        """
        - Лидеры по числу побед
        """
        con, cur = library.get.con_cur()
        guild_id = library.get.guild_id(ctx)
        select = Const.selects.USWins
        cur.execute(select, (guild_id, count))
        records = cur.fetchall()
        embed = Embed(
            title=f"Таблица лидеров",
            color=config.info
        )
        value = ""
        for i, record in enumerate(records):
            value += f"{i + 1}. {library.get.mention(record.id)} — {record.win}\n"
        embed.add_field(
            name=f"Топ {count} игроков по числу побед",
            value=value
        )
        await ctx.send(embed=embed)

    @top.command(name="points")
    async def top_points(self, ctx, count=10):
        """
        - Лидеры по заработанным очкам
        """
        con, cur = library.get.con_cur()
        guild_id = library.get.guild_id(ctx)
        select = Const.selects.USPoints
        cur.execute(select, (guild_id, count, ))
        records = cur.fetchall()
        embed = Embed(
            title=f"Таблица лидеров",
            color=config.info
        )
        value = ""
        # max_indent = len(max([self.bot.get_user(x.id).name for x in records])) + 1  # получать макс длину имени
        for i, record in enumerate(records):
            value += f"{i+1}. {library.get.mention(record.id)} — {record.points}\n"
        embed.add_field(
            name=f"Топ {count} игроков по числу баллов",
            value=value
        )
        await ctx.send(embed=embed)

    @top.command(name="remove")
    @commands.check_any(commands.has_role(703884580041785344),  # Создатель
                        commands.has_role(703884637755408466),  # Админ
                        commands.has_role(711230509967212564),  # Старший модер
                        commands.has_role(711230315540250624),  # Модер
                        commands.has_role(946480695218429952),  # Samuro_dev
                        commands.has_role(880865537058545686),  # test
                        )
    async def points_remove(self, ctx, user: Member, count=0):
        con, cur = library.get.con_cur()
        guild_id = library.get.guild_id(ctx)
        select = Const.selects.USIdGuild
        cur.execute(select, (user.id, guild_id))
        record = cur.fetchone()
        stats = library.get.stats(record)
        if stats.points < count:
            await ctx.send(f"Недостаточно баллов\n"
                           f"Баллов у {library.get.mention(stats.id)}: {stats.points}")
        else:
            stats.points -= count
            update = Const.updates.USPoints
            cur.execute(update, (stats.points, stats.id, stats.guild_id))
            pl.commit(con)
            await ctx.send(f"Баллы успешно сняты\n"
                           f"Осталось баллов: {stats.points}")

    @points_remove.error
    @top_mmr.error
    async def points_handler(self, ctx, error):
        error = getattr(error, 'original', error)  # получаем пользовательские ошибки
        print(error)
        if isinstance(error, commands.errors.MissingRole):
            await ctx.send("Недостаточно прав для выполнения команды")
        elif isinstance(error, exceptions.WrongLeague):
            await ctx.send("Выберите корректную лигу")


def setup(bot):
    bot.add_cog(Stats(bot))